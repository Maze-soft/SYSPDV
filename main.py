import tkinter as tk
import os
from tkinter import ttk
import datetime
from time import strftime
import mysql.connector as sql
import pymysql
from PIL import Image, ImageTk
from tkinter import filedialog
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from impressora import Conector, AccionBarcodeJan13, AlineacionCentro
import matplotlib.pyplot as plt



pymysql.install_as_MySQLdb()


da =  datetime.date.today()

bg_ = ("#282828")


class syspdv(tk.Tk):

    def __init__(self, *args, **kwargs):
        
        tk.Tk.__init__(self, *args, **kwargs)
        container = tk.Frame(self)

        container.pack(side="top", fill="both", expand = True)

        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        for F in (Inicio, Caixa, Produtos, Vendas, Estoque):

            frame = F(container, self, bg=bg_)

            self.frames[F] = frame

            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(Inicio)

    def show_frame(self, cont):

        frame = self.frames[cont]
        frame.tkraise()

        
class Inicio(tk.Frame):

    def __init__(self, parent, controller,bg=None):
        tk.Frame.__init__(self,parent, bg=bg)
        
        #filename = ImageTk.PhotoImage(Image.open('sys.png'))
        #background_label = tk.Label(self, image=filename, width=450, height=300)
        #background_label.place(x=320, y=80)
        

        button = tk.Button(self, text="C A I X A",
                            command=lambda: controller.show_frame(Caixa)
                           ,width=80, height=10, fg="white",bg="#ffae00"
                           ,font="arial 8 bold")
        button.place(x=30,y=100)

        button2 = tk.Button(self, text="P R O D U T O S",
                            command=lambda: controller.show_frame(Produtos),
                            width=80, height=10,fg="white",bg="#008cff"
                            ,font="arial 8 bold")
        button2.place(x=750, y=100)
        
        button3 = tk.Button(self, text="V E N D A S",
                            command=lambda: controller.show_frame(Vendas),
                            width=80, height=10, fg="white", bg="#00ff59"
                            ,font="arial 8 bold")
        button3.place(x=30, y=400)

        button4 = tk.Button(self, text="E S T O Q U E",
                            command=lambda: controller.show_frame(Estoque)
                            ,width=80, height=10, fg="white", bg="#ff94b8"
                            ,font="arial 8 bold")
        button4.place(x=750,y=400)


        hora = tk.Label(self, text="v1.5.7 - BETA . SYSPDV SOFTWARE CORPORATION.", fg="white", font="arial 16 bold"
           ,bg="#282828").place(x=30,y=600)

        def tempo():
            tempos = strftime('%H:%M:%S %p')
            horas.config(text=f"HORA: {tempos}")
            horas.after(1000,tempo)

        horas = tk.Label(self, fg="white", font="arial 16 bold"
           ,bg="#282828")
        horas.place(x=750,y=600)

        tempo()
        
        data = tk.Label(self, text=f"DATA: {da}", fg="white", font="arial 16 bold"
           ,bg="#282828").place(x=1050,y=600)
        
i=10000
total=0

class Caixa(tk.Frame):

    def __init__(self, parent, controller,bg=None):
        tk.Frame.__init__(self, parent,bg=bg)

        #ADC MYSQL

        def deletar():
            app.destroy()
            os.popen("main.py")
        
        
        def adc():
            global total

            qt = qt_text.get()
            price = pre_text.get()
            


            preco_prod = float(qt) * float(price)
            total += float(qt) * float(price)


            tota["text"] = "R$",total
            nome = n_text.get()
            produ["text"] = nome, "-", qt, "unidade(s)", "por R$", price , "cada"
            
            liss = lista.insert(i, n_text.get() + "-"  + "R$" + str(preco_prod), str(qt), "unidade(s)", "por R$", str(price))

        def troco():
            global total
            pago = pa_text.get()

            qt = qt_text.get()
            price = pre_text.get()


            tr = float(pago) - float(total)

            trocos["text"] = "R$",tr

        def concluir():
            global total

            dia = da.day
            mes=da.month
    
            con = sql.connect(host="localhost", user="root",
                                    password="META100Kk#", database="mercado")
            cur = con.cursor()
            query = "INSERT INTO vendas(id, valor, mês, dia) VALUES(null, %s, %s, %s)"
            val = ((total,mes, dia))
            cur.execute(query, val)
            con.commit()
            con.close()

            '''

            print("===================================")
            print("============= LOJA ================")
            print("========= SEJA BEM-VINDO ==========")
            '''
            
            price = pre_text.get()
            qt = qt_text.get()
            preco_prod = float(qt) * float(price)
            
            pago = pa_text.get()

            get = lista.get(0,tk.END)
            pr = [str(i) for i in get]
            prod_list = str("".join(pr))

            #for con_item in get_content:
                #con1, con2 = con_item.split('PHP')
                #rec_content = f'{con1:<40}costs PHP{con2:<8}'
                
            
            
            tr = float(pago) - float(total)
            #print(cod_text.get(), preco_prod)
            '''
            print("=========== VALOREES ==============")
            print("Valor Total:R$", total)
            print("Troco:R$", tr)
            print("===================================")
            '''
            
            impresoras = Conector.obtenerImpresoras()
            print(f"As impressoras são: {impresoras}")

            c = Conector()
            c.textoConAcentos("====================\n")
            c.establecerEnfatizado(1)
            c.texto("Loja\supermercado\mercado\n")
            c.establecerEnfatizado(0)
            c.texto("SEJA BEM-VINDO(a)\n")
            c.establecerEnfatizado(1)
            c.texto("=====================\n")
            c.establecerTamanioFuente(1, 1)
            c.establecerJustificacion(AlineacionCentro)
            c.texto("prod_list\n")
            #c.texto("Código de barras:\n")
            c.establecerEnfatizado(1)
            c.texto("=======VALORES=========\n")
            c.establecerEnfatizado(1)
            c.texto("Valor TOTAL:R$",total,"\n")
            c.establecerEnfatizado(1)
            c.texto("TROCO:R$",tr,"\n")
            c.establecerEnfatizado(1)
            c.texto("=====================\n")
            c.feed(5)
            c.cortar()
            c.abrirCajon()
            print("Imprimindo...")
            # Recuerda cambiar por el nombre de tu impresora
            respuesta = c.imprimirEn("ZJ-58")
            if respuesta == True:
                print("Impressao concluida")
            else:
                print(f"Error. A mensagem diz: {respuesta}")
            
        def barra():
            global total

            while True:

                print("Lendo Codigo de barra")

                codigo = int(input("Digite seu codigo: "))

                if codigo == 0:
                    break
                
                con = sql.connect(host="localhost", user="root",
                              password="META100Kk#", database="mercado")
                seleciona = "SELECT codigo FROM produtos WHERE codigo ='{}'".format(codigo)
                cur = con.cursor()
                cur.execute(seleciona)
                resultado = cur.fetchall()
                
                qnt = int(input("Quantidade: "))
                if len(resultado)!=0:
                    print(True)
                    #lb["text"] = codigo
                        
                    price = "SELECT Preco FROM produtos WHERE codigo={}".format(codigo)
                    name = "SELECT Nome FROM produtos WHERE codigo={}".format(codigo)
                    kur = con.cursor()
                    kur.execute(name)
                    nm = kur.fetchone()
                    nms = [str(i) for i in nm]
                    resultados_nome = str("".join(nms))
                
                    
                    cur = con.cursor()
                    cur.execute(price)
                    price_result = cur.fetchone()
                    num = [str(i) for i in price_result]
                    resultadoss = str("".join(num))
                    
                else:
                    print(False)

                


                pri = resultadoss
                    
                try:
                    qt = int(qnt)
                except:
                    qt = 0.0
                #print(int(pri) * qt)


                preco_prod = qt * int(pri)
                print(preco_prod)

                total += qt * int(pri)
                tota["text"] = "R$",total
                #nome = n_text.get()
                produ["text"] = resultados_nome, "-", qnt, "unidade(s)", "por R$", resultadoss , "cada"
                
                liss = lista.insert(i, resultados_nome +"-"  + "R$" + str(preco_prod),"- R$", str(resultados))
                




        label = tk.Label(self, text="C A I X A", font="arial 26 bold",bg="#282828"
                         ,fg="white")
        label.place(x=10,y=10)

        produ = tk.Label(self, text="{PRODUTO_NOME}", font="arial 26 bold",bg="#282828"
                         ,fg="white")
        produ.place(x=330,y=10)

        '''

        label = tk.Label(self, text="PAGAMENTO", font="arial 18 bold",bg="#282828"
                         ,fg="white")
        label.place(x=1170,y=130)

        var = tk.StringVar()
    
        din = tk.Radiobutton(self, text="D I N H E I R O", variable=var, value=1,
                          bg="#282828", fg="white")
        din.place(x=1170,y=200)
        pix = tk.Radiobutton(self, text="P I X", variable=var, value=2,
                          bg="#282828", fg="white",command=pix)
        pix.place(x=1170,y=220)
        deb = tk.Radiobutton(self, text="D E B I T O", variable=var, value=3,
                          bg="#282828", fg="white",command=debito)
        deb.place(x=1170,y=240)
        cred = tk.Radiobutton(self, text="C R E D I T O", variable=var, value=4,
                          bg="#282828", fg="white",command=credi)
        cred.place(x=1170,y=260)

        '''

        label = tk.Label(self, text="T O T A L", font="arial 18 bold",bg="#282828"
                         ,fg="white")
        label.place(x=1170,y=130)

        tota = tk.Label(self, text="R$00,00", font="arial 26 bold",bg="#282828"
                         ,fg="white")
        tota.place(x=1170,y=170)

        label = tk.Label(self, text="T R O C O", font="arial 18 bold",bg="#282828"
                         ,fg="white")
        label.place(x=1170,y=220)

        trocos = tk.Label(self, text="R$00,00", font="arial 26 bold",bg="#282828"
                         ,fg="white")
        trocos.place(x=1170,y=260)



        button1 = tk.Button(self, text="<",
                            command=lambda: controller.show_frame(Inicio),
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=12, height=2)
        button1.place(x=10, y=650)

        def calcu():
            os.startfile("C:\Windows\System32\calc.exe")

        
        

        cal = tk.Button(self, text="C A L C U L A D O R A",
                            command=calcu,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=18, height=2)
        cal.place(x=1170, y=650)

        cod = tk.Button(self, text="C O D \ B A R R A",
                            command=barra,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=18, height=2)
        cod.place(x=1170, y=10)

        button1 = tk.Button(self, text="C O N C L U I R",
                            command=concluir,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=45, height=2)
        button1.place(x=650, y=650)

        button1 = tk.Button(self, text="L I M P A R",
                            command=deletar,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=45, height=2)
        button1.place(x=200, y=650)

        n = tk.Label(self, text="N O M E / C O D",bg="#282828"
                         ,fg="white", font="Arial 10 ").place(x=10, y=100)
        n_text= tk.Entry(self, width=25, font="arial 16")
        n_text.place(x=10,y=140)
        qt = tk.Label(self, text="Q U A N T I D A D E",bg="#282828"
                         ,fg="white", font="Arial 10 ").place(x=10, y=180)
        qt_text= tk.Entry(self, width=25, font="arial 16")
        qt_text.place(x=10,y=220)
        
        pre = tk.Label(self, text="P R E Ç O",bg="#282828"
                         ,fg="white", font="Arial 10 ").place(x=10, y=260)
        pre_text= tk.Entry(self, width=25, font="arial 16")
        pre_text.place(x=10,y=300)

        button1 = tk.Button(self, text="A D I C I O N A R",
                            command=adc,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=30, height=2)
        button1.place(x=10, y=360)

        pre = tk.Label(self, text="V A L \ P A G O",bg="#282828"
                         ,fg="white", font="Arial 10 ").place(x=10, y=420)
        pa_text= tk.Entry(self, width=25, font="arial 16")
        pa_text.place(x=10,y=460)

        confer = tk.Button(self, text="C O N F E R I R",
                            command=troco,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=30, height=2)
        confer.place(x=10, y=520)

        lista = tk.Listbox(self, height = 24,width = 85,bg = "white",activestyle = 'dotbox',
                  font = "Arial 12")

        lista.place(x=330,y=130)



class Produtos(tk.Frame):

    def __init__(self, parent, controller,bg=None):
        tk.Frame.__init__(self, parent,bg=bg)
        def criar_prod():
            name = n_text.get()
            cod = code_text.get()
            pr = pre_text.get()
            qt= qt_text.get()
            
            con = sql.connect(host="localhost", user="root",
                              password="META100Kk#", database="mercado")
            cur = con.cursor()
            query = "INSERT INTO produtos(id, nome, preco, codigo, quantidade) VALUES(null, %s, %s, %s, %s)"
            val = (name, pr, cod, qt)
            cur.execute(query, val)
            con.commit()
            con.close()

            nome["text"] = name
            codi["text"] = cod
            price["text"] = pr
            qtd["text"] = qt

        def atua():
            app.destroy()
            os.popen("main.py")



        def remover():
            
            
            selected = table.selection()
            
            
            if selected:
                x = selected[0]
                item = table.item(x)["values"][0]
                table.delete(x)

                con = sql.connect(host="localhost", user="root",
                                      password="META100Kk#", database="mercado")
                seleciona = "DELETE FROM produtos WHERE id ='{}'".format(item)
                cur = con.cursor()
                cur.execute(seleciona)
                    
                #command = 'DELETE FROM produtos WHERE id=%s'
                #cur.execute(command, (item,))
            
                con.commit()
                
        def excel():
            #file = filedialog.asksaveasfilename(title="Select file", "produtos.xlsx",filetypes[("Excel file", "*.xlsx")])
             workbook = load_workbook(filename='produtos.xlsx')

             for worksheets in workbook.sheetnames: 
                 sheet=workbook[worksheets]
                 sheet.delete_rows(idx=2, amount=15)
                 for row_id in table.get_children():
                    row = table.item(row_id)['values']
                    sheet.append(row)
                 workbook.save(filename='produtos.xlsx')   
            

            
            
        label = tk.Label(self, text="P R O D U T O S", font="arial 26 bold",bg="#282828"
                         ,fg="white")
        label.place(x=10,y=10)

        label = tk.Label(self, text="L I S T A", font="arial 10 bold",bg="#282828"
                         ,fg="white")
        label.place(x=10,y=50)

        criar = tk.Label(self, text="C R I A R",bg="#282828"
                         ,fg="white", font="Arial 20 bold").place(x=10, y=120)
    
        n = tk.Label(self, text="N O M E",bg="#282828"
                         ,fg="white", font="Arial 10 ").place(x=10, y=180)
        n_text= tk.Entry(self, width=25, font="arial 16")
        n_text.place(x=10,y=220)
        code = tk.Label(self, text="C O D I G O",bg="#282828"
                         ,fg="white", font="Arial 10 ").place(x=10, y=260)
        code_text= tk.Entry(self, width=25, font="arial 16")
        code_text.place(x=10,y=300)
        qt = tk.Label(self, text="Q U A N T I D A D E",bg="#282828"
                         ,fg="white", font="Arial 10 ").place(x=10, y=340)
        qt_text= tk.Entry(self, width=25, font="arial 16")
        qt_text.place(x=10,y=380)
        
        pre = tk.Label(self, text="P R E Ç O",bg="#282828"
                         ,fg="white", font="Arial 10 ").place(x=10, y=420)
        pre_text= tk.Entry(self, width=25, font="arial 16")
        pre_text.place(x=10,y=460)

        button1 = tk.Button(self, text="A D I C I O N A R",
                            command=criar_prod,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=30, height=2)
        button1.place(x=10, y=500)

        button1 = tk.Button(self, text="A T U A L I Z A R",
                            command=atua,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=45, height=2)
        button1.place(x=650, y=650)

        button1 = tk.Button(self, text="E X C L U I R",
                            command=remover,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=45, height=2)
        button1.place(x=200, y=650)
        

        
        button1 = tk.Button(self, text="<",
                            command=lambda: controller.show_frame(Inicio),
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=12, height=2)
        button1.place(x=10, y=650)

        label = tk.Label(self,height=12,bg="#282828")
        label.pack()

        tabela = tk.Frame(self)
        tabela.pack(anchor="center")

        table = ttk.Treeview(tabela,height=20)

        table['show'] = 'headings'

        table["columns"] = ("id", "Nome", "Preço", "Codigo", "Quantidade")
        table.column("#0", width=0)
        table.column("id",  width=10)
        table.column("Nome",  width=150)
        table.column("Preço", width=150)
        table.column("Codigo",  width=150)
        table.column("Quantidade",  width=150)

        table.heading("#0", text="")
        table.heading("id", text="ID")
        table.heading("Nome", text="Nome")
        table.heading("Preço", text="Preço")
        table.heading("Codigo", text="Codigo")
        table.heading("Quantidade", text="Quantidade")
    
        
        con = sql.connect(host="localhost", user="root",
                  password="META100Kk#", database="mercado")
        cur = con.cursor()
        cur.execute("SELECT * FROM produtos")
        

        ########
        i = 0
        for ro in cur:
            
            table.insert("",i, text="", values=(ro[0], ro[1], ro[2], ro[3], ro[4]))
            i = i + 1
    
        table.pack()

        ex = tk.Button(self, text="E X C E L",
                            command=excel,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=18, height=2)
        ex.place(x=1170, y=10)

        label = tk.Label(self, text="N O M E", font="arial 18 bold",bg="#282828"
                         ,fg="white")
        label.place(x=1100,y=200)

        nome = tk.Label(self, text="{nome}", font="arial 26 bold",bg="#282828"
                         ,fg="white")
        nome.place(x=1100,y=240)

        label = tk.Label(self, text="C O D I G O", font="arial 18 bold",bg="#282828"
                         ,fg="white")
        label.place(x=1100,y=300)

        codi = tk.Label(self, text="{code}", font="arial 26 bold",bg="#282828"
                         ,fg="white")
        codi.place(x=1100,y=340)

        label = tk.Label(self, text="P R E Ç O", font="arial 18 bold",bg="#282828"
                         ,fg="white")
        label.place(x=1100,y=400)

        price = tk.Label(self, text="{price}", font="arial 26 bold",bg="#282828"
                         ,fg="white")
        price.place(x=1100,y=440)

        label = tk.Label(self, text="Q U A N T I D A D E", font="arial 18 bold",bg="#282828"
                         ,fg="white")
        label.place(x=1100,y=500)

        qtd = tk.Label(self, text="{qtd}", font="arial 26 bold",bg="#282828"
                         ,fg="white")
        qtd.place(x=1100,y=540)

        

class Vendas(tk.Frame):

    def __init__(self, parent, controller,bg=None):
        tk.Frame.__init__(self, parent,bg=bg)
        
            
        def filtrar():
            dia = value_inside.get()
            mes = value_inside2.get()
            if dia and mes:
                table.delete(*table.get_children())
                conn = sql.connect(host="localhost", user="root",
                  password="META100Kk#", database="mercado")
                cursor = conn.cursor()
                cursor.execute("SELECT id, valor,mês,dia FROM vendas WHERE dia ='{}' and mês ='{}'".format(dia,mes))
                fetch = cursor.fetchall()
                for data in fetch:
                    table.insert('', 'end', values=(data))
                cursor.close()
                conn.close()
            

            
        def excel():
            #file = filedialog.asksaveasfilename(title="Select file","vendas.xlsx",filetypes[("Excel file", "*.xlsx")])
            workbook = load_workbook(filename='vendas.xlsx')

            for worksheets in workbook.sheetnames: 
                sheet=workbook[worksheets]
                sheet.delete_rows(idx=2, amount=15)
                for row_id in table.get_children():
                    row = table.item(row_id)['values']
                    sheet.append(row)
                workbook.save(filename='vendas.xlsx')
            
        label = tk.Label(self, text="V E N D A S ", font="arial 26 bold",bg="#282828"
                         ,fg="white")
        label.place(x=10,y=10)

        label = tk.Label(self, text="L I S T A", font="arial 10 bold",bg="#282828"
                         ,fg="white")
        label.place(x=10,y=50)

        criar = tk.Label(self, text="F I L T R A R  V E N D A S",bg="#282828"
                         ,fg="white", font="Arial 20 bold").place(x=10, y=120)

        button1 = tk.Button(self, text="<",
                            command=lambda: controller.show_frame(Inicio),
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=12, height=2)
        button1.place(x=10,y=650)

        label = tk.Label(self,height=12,bg="#282828")
        label.pack()

        tabela = tk.Frame(self)
        tabela.pack(anchor="center")

        table = ttk.Treeview(tabela,height=20)

        table['show'] = 'headings'

        table["columns"] = ("id", "Valor", "Mês", "Dia")
        table.column("#0", width=0)
        table.column("id",  width=150)
        table.column("Valor",  width=150)
        table.column("Mês", width=150)
        table.column("Dia", width=150)
        

        table.heading("#0", text="")
        table.heading("id", text="ID")
        table.heading("Valor", text="Valor")
        table.heading("Mês", text="Mês")
        table.heading("Dia", text="Dia")
        
        con = sql.connect(host="localhost", user="root",
                  password="META100Kk#", database="mercado")
        cur = con.cursor()
        cur.execute("SELECT * FROM vendas")
        

        ########
        i = 0
        for ro in cur:
            table.insert("", i, text="", values=(ro[0], ro[1], ro[2], ro[3]))
            i = i + 1
         
        table.pack()

        cod = tk.Button(self, text="E X C E L",
                            command=excel,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=18, height=2)
        cod.place(x=1170, y=10)

        options_dia = ["01", "02", "03", "04",
                        "05", "06", "07", "08",
                        "09", "10", "11", "12",
                        "13", "14", "15", "16",
                        "17", "18", "19", "20",
                        "21", "22", "23", "24", "25", "26",
                        "27", "28", "29", "30", "31"]
  

        value_inside = tk.StringVar(self)
  
        value_inside.set("Selecione um dia")
  
        question_dia = tk.OptionMenu(self, value_inside, *options_dia)
        question_dia.config(bg="#282828",fg="white")
        question_dia.place(x=10, y=185)

        options_mes = ["01","02","03","04","05","06","07","08","09"
                        ,"10","11","12"]
  

        value_inside2 = tk.StringVar(self)
  
        value_inside2.set("Selecione um mês")
  
        question_mes = tk.OptionMenu(self, value_inside2, *options_mes)
        question_mes.config(bg="#282828",fg="white")
        question_mes.place(x=10, y=255)

        button1 = tk.Button(self, text="F I L T R A R",
                            command=filtrar,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=30, height=2)
        button1.place(x=10, y=320)

        

    
ido=0

class Estoque(tk.Frame):

    def __init__(self, parent, controller,bg=None):
        tk.Frame.__init__(self, parent,bg=bg)
        label = tk.Label(self, text="E S T O Q U E", font="arial 26 bold",bg="#282828"
                         ,fg="white")
        label.place(x=10,y=10)
        def atua():
            app.destroy()
            os.popen("main.py")

        def excel():
            workbook = load_workbook(filename='estoque.xlsx')

            for worksheets in workbook.sheetnames: 
                sheet=workbook[worksheets]
                sheet.delete_rows(idx=2, amount=15)
                for row_id in table.get_children():
                    row = table.item(row_id)['values']
                    sheet.append(row)
                workbook.save(filename='estoque.xlsx')
        
        def edit():

            nome = n_text.get()
            code = code_text.get()
            qtd=qt_text.get()
            pre=pre_text.get()
            
            selected = table.selection()
            
            
            if selected:
                x = selected[0]
                item = table.item(x)["values"][0]
                table.delete(x)

                con = sql.connect(host="localhost", user="root",
                                      password="META100Kk#", database="mercado")
                seleciona = "UPDATE produtos SET Nome='{}', Preco='{}', codigo='{}', quantidade='{}' WHERE id ='{}'".format(nome, pre, code, qtd, item)
                cur = con.cursor()
                cur.execute(seleciona)

                con.commit()
                
            #UPDATE books SET title = %s WHERE id = %s
            #selected_item = table.selection()[0]
            #table.item(selected_item, text="blub", values=(nome, code, qtd, pre))

        criar = tk.Label(self, text="S E L E C I O N A R",bg="#282828"
                         ,fg="white", font="Arial 20 bold").place(x=10, y=120)

        label = tk.Label(self, text="L I S T A", font="arial 10 bold",bg="#282828"
                         ,fg="white")
        label.place(x=10,y=50)

        button1 = tk.Button(self, text="<",
                            command=lambda: controller.show_frame(Inicio),
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=12, height=2)
        button1.place(x=10, y=650)

        label = tk.Label(self,height=12,bg="#282828")
        label.pack()

        tabela = tk.Frame(self)
        tabela.pack(anchor="center")

        table = ttk.Treeview(tabela,height=20)

        table['show'] = 'headings'

        table["columns"] = ("id", "Nome", "Preço", "Codigo", "Quantidade")
        table.column("#0", width=0)
        table.column("id",  width=10)
        table.column("Nome",  width=150)
        table.column("Preço", width=150)
        table.column("Codigo",  width=150)
        table.column("Quantidade",  width=150)

        table.heading("#0", text="")
        table.heading("id", text="ID")
        table.heading("Nome", text="Nome")
        table.heading("Preço", text="Preço")
        table.heading("Codigo", text="Codigo")
        table.heading("Quantidade", text="Quantidade")
        con = sql.connect(host="localhost", user="root",
                  password="META100Kk#", database="mercado")
        cur = con.cursor()
        cur.execute("SELECT * FROM produtos")
        

        ########
        i = 0
        for ro in cur:
            table.insert("", i, text="", values=(ro[0], ro[1], ro[2], ro[3], ro[4]))
            i = i + 1
    
        table.pack()

        cod = tk.Button(self, text="E X C E L",
                        command=excel,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=18, height=2)
        cod.place(x=1170, y=10)

        n = tk.Label(self, text="N O M E",bg="#282828"
                         ,fg="white", font="Arial 10 ").place(x=10, y=180)
        n_text= tk.Entry(self, width=25, font="arial 16")
        n_text.place(x=10,y=220)
        code = tk.Label(self, text="C O D I G O",bg="#282828"
                         ,fg="white", font="Arial 10 ").place(x=10, y=260)
        code_text= tk.Entry(self, width=25, font="arial 16")
        code_text.place(x=10,y=300)
        qt = tk.Label(self, text="Q U A N T I D A D E",bg="#282828"
                         ,fg="white", font="Arial 10 ").place(x=10, y=340)
        qt_text= tk.Entry(self, width=25, font="arial 16")
        qt_text.place(x=10,y=380)
        
        pre = tk.Label(self, text="P R E Ç O",bg="#282828"
                         ,fg="white", font="Arial 10 ").place(x=10, y=420)
        pre_text= tk.Entry(self, width=25, font="arial 16")
        pre_text.place(x=10,y=460)

        button1 = tk.Button(self, text="E D I T A R",
                            command=edit,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=30, height=2)
        button1.place(x=10, y=500)

        label = tk.Label(self, text="N O M E", font="arial 18 bold",bg="#282828"
                         ,fg="white")
        label.place(x=1100,y=200)

        label = tk.Label(self, text="{nome}", font="arial 26 bold",bg="#282828"
                         ,fg="white")
        label.place(x=1100,y=240)

        label = tk.Label(self, text="Q U A N T I D A D E", font="arial 18 bold",bg="#282828"
                         ,fg="white")
        label.place(x=1100,y=300)

        label = tk.Label(self, text="{qtd}", font="arial 26 bold",bg="#282828"
                         ,fg="white")
        label.place(x=1100,y=340)

        cal = tk.Button(self, text="A T U A L I Z A R",
                            command=atua,
                            bg="#282828", fg="white", font="arial 12 bold",
                            width=18, height=2)
        cal.place(x=1170, y=650)

        
    

app = syspdv()
app.geometry("1000x1000")
app.title("SOFTWARE DE GESTÃO DE EMPRESAS - SYSPDV")
app.mainloop()
